from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import os
from dotenv import load_dotenv
load_dotenv()

# Initialize FastAPI
app = FastAPI(title="Mamameal API")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure Gemini
api_key = os.getenv("GOOGLE_API_KEY")
if not api_key:
    print("Warning: GOOGLE_API_KEY not found in environment variables.")


@app.get("/")
def read_root():
    return {"message": "Mamameal API is running"}

@app.get("/health")
def health_check():
    return {"status": "ok"}

from api.seal_utils import generate_seal_data, create_seal_excel
import base64

@app.post("/api/seal")
async def create_seal(file: UploadFile = File(...)):
    try:
        content = await file.read()
        blocks = generate_seal_data(content, api_key=api_key)
        excel_io = create_seal_excel(blocks)
        
        # Return as base64 for now (or streaming response)
        b64_str = base64.b64encode(excel_io.getvalue()).decode()
        return {
            "filename": f"{file.filename.replace('.pdf', '')}_seal.xlsx",
            "file_data": b64_str,
            "blocks": blocks
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Order/Invoice Processing ---
from api.pdf_utils import (
    safe_write_df, match_bento_data, paste_dataframe_to_sheet,
    extract_detailed_client_info_from_pdf
)
# Note: Legacy extraction functions removed/unused in favor of AI
from api.ai_processor import process_order_pdf_with_ai
from api.master_utils import load_master_csv, save_master_file
from openpyxl import load_workbook
import io
import pandas as pd

# Assets directory - configurable for Railway Volume
# In Railway, set ASSETS_DIR env var to "/app/api/assets" (Volume mount path)
ASSETS_DIR = os.getenv("ASSETS_DIR", os.path.join(os.path.dirname(__file__), 'api', 'assets'))

@app.post("/api/order-invoice")
async def process_order(file: UploadFile = File(...)):
    try:
        pdf_bytes = await file.read()
        
        # Load Masters
        df_product_master, _ = load_master_csv(ASSETS_DIR, "商品マスタ")
        df_customer_master, _ = load_master_csv(ASSETS_DIR, "得意先マスタ")
        
        # --- AI Extraction ---
        if not api_key:
             raise HTTPException(status_code=500, detail="API Key not configured for AI processing")

        ai_result = process_order_pdf_with_ai(pdf_bytes, api_key)
        bento_header_names = ai_result.get('bento_headers', [])
        
        # 1. Process Clients (Legacy Layout Extraction)
        # Reverting to original 'mamameal-next' logic as requested by user.
        # This uses 'extract_text_with_layout' to robustly find rows.
        df_client_sheet = None
        client_data_legacy = extract_detailed_client_info_from_pdf(io.BytesIO(pdf_bytes))
        
        num_bento_cols = len(bento_header_names) if bento_header_names else 5 # Default to 5 to be safe if no AI headers?
        
        # Create lookup dictionary: 得意先名(B列) -> 得意先名略称(D列)
        customer_name_map = {}
        if not df_customer_master.empty:
            df_customer_master.columns = df_customer_master.columns.str.strip()
            # B列=得意先名, D列=得意先名略称 (column indices 1 and 3)
            col_names = list(df_customer_master.columns)
            if len(col_names) >= 4:
                internal_name_col = col_names[1]  # B列: 得意先名
                customer_name_col = col_names[3]  # D列: 得意先名略称
                for _, row in df_customer_master.iterrows():
                    internal_name = str(row[internal_name_col]).strip()
                    customer_name = str(row[customer_name_col]).strip()
                    if internal_name and customer_name:
                        customer_name_map[internal_name] = customer_name
        
        client_rows = []
        for info in client_data_legacy:
            s_list = info.get('student_meals', [])
            t_list = info.get('teacher_meals', [])
            
            internal_client_name = info['client_name']
            # Lookup customer-facing name from master
            customer_facing_name = customer_name_map.get(internal_client_name, '')
            
            # Add client_id (A列), client_name (B列), customer_facing_name (C列)
            row = {
                'クライアントID': info.get('client_id', ''),
                'クライアント名': internal_client_name,
                'クライアント名（顧客向け）': customer_facing_name
            }
            
            # Dynamic Columns: Student 1..N (D列から開始)
            for i in range(num_bento_cols):
                val = s_list[i] if i < len(s_list) else ''
                row[f's_{i}'] = val
                
            # Dynamic Columns: Teacher 1..N
            for i in range(num_bento_cols):
                val = t_list[i] if i < len(t_list) else ''
                row[f't_{i}'] = val
            
            client_rows.append(row)
            
        if client_rows:
            df_client_sheet = pd.DataFrame(client_rows)
            # Enforce column order: ID, Client, CustomerName, S...S, T...T
            cols = ['クライアントID', 'クライアント名', 'クライアント名（顧客向け）'] + [f's_{i}' for i in range(num_bento_cols)] + [f't_{i}' for i in range(num_bento_cols)]
            # Ensure only existing columns are selected (in case df was created differently?)
            # creating from list of dicts creates all keys.
            df_client_sheet = df_client_sheet[cols]

        # 2. Process Bentos
        df_bento_sheet = None
        # Use 'bento_headers' strings extracted by AI
        bento_names = ai_result.get('bento_headers', [])
        
        if bento_names:
            matched_data = []
            for b_name in bento_names:
                search_key = b_name
                # User requested grouping logic:
                # "Charaben..." -> Search "キャラ"
                # "Red..." -> Search "赤"
                if "キャラ弁" in b_name:
                    search_key = "キャラ"
                elif b_name.startswith("赤 ") or b_name == "赤": # Match "赤 飯あり..."
                    search_key = "赤"
                    
                # Match against Product Master (returns [[Name, Box, Price, Type]])
                # We use search_key to find the row, but keep b_name for the Excel Name column
                match_res = match_bento_data([search_key], df_product_master)
                if match_res:
                    m_row = match_res[0]
                    # [Original_AI_Name, Box, Price, Type]
                    matched_data.append([b_name, m_row[1], m_row[2], m_row[3]])
            
            df_bento_sheet = pd.DataFrame(matched_data, columns=['商品予定名', 'パン箱入数', '売価単価', '弁当区分'])

        # 3. Paste Sheet (Legacy) -> Empty
        df_paste_sheet = pd.DataFrame() 


        # 4. Generate Files
        template_path = os.path.join(ASSETS_DIR, "template.xlsm")
        nouhinsyo_path = os.path.join(ASSETS_DIR, "nouhinsyo.xlsx")
        
        if not os.path.exists(template_path) or not os.path.exists(nouhinsyo_path):
             raise HTTPException(status_code=500, detail="Template files not found")

        # --- Generate Template (Shuushussho) ---
        template_wb = load_workbook(template_path, keep_vba=True)
        
        # Paste Masters
        if not df_product_master.empty and "商品マスタ" in template_wb.sheetnames:
            ws = template_wb["商品マスタ"]
            if ws.max_row > 0: ws.delete_rows(1, ws.max_row)
            paste_dataframe_to_sheet(ws, df_product_master)
            
        if not df_customer_master.empty and "得意先マスタ" in template_wb.sheetnames:
             ws = template_wb["得意先マスタ"]
             if ws.max_row > 0: ws.delete_rows(1, ws.max_row)
             paste_dataframe_to_sheet(ws, df_customer_master)

        # Write Data
        ws_paste = template_wb["貼り付け用"]
        for r_idx, row in df_paste_sheet.iterrows():
            for c_idx, value in enumerate(row):
                ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
        
        if df_bento_sheet is not None and "注文弁当の抽出" in template_wb.sheetnames:
            safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet)
            
        if df_client_sheet is not None and "クライアント抽出" in template_wb.sheetnames:
            ws_client = template_wb["クライアント抽出"]
            # Write DATA starting at Row 2 (leaving Row 1 for headers)
            # This ensures we don't overwrite our new dynamic headers
            safe_write_df(ws_client, df_client_sheet, start_row=2)
            
            # --- Dynamic Header Injection (Explicit Write) ---
            # Manually write headers to Row 1 because safe_write_df reads values only
            # The structure is: ID | Client | CustomerName | Student_Cols... | Teacher_Cols...
            ws_client.cell(row=1, column=1, value='クライアントID')
            ws_client.cell(row=1, column=2, value='クライアント名')
            ws_client.cell(row=1, column=3, value='クライアント名（顧客向け）')
            
            if bento_header_names:
                num_cols = len(bento_header_names)
                for i in range(num_cols):
                    b_name = bento_header_names[i]
                    
                    # Student Header (Starts at Col 4 = D列)
                    ws_client.cell(row=1, column=4+i, value=f"{b_name}\n(園児)")
                    
                    # Teacher Header (Starts after Student Block)
                    ws_client.cell(row=1, column=4+num_cols+i, value=f"{b_name}\n(先生)")
            # -------------------------------------------------
            
        out_template = io.BytesIO()
        template_wb.save(out_template)
        b64_template = base64.b64encode(out_template.getvalue()).decode()

        # --- Generate Nouhinsyo ---
        nouhinsyo_wb = load_workbook(nouhinsyo_path)
        
        if not df_customer_master.empty and "得意先マスタ" in nouhinsyo_wb.sheetnames:
             ws = nouhinsyo_wb["得意先マスタ"]
             if ws.max_row > 0: ws.delete_rows(1, ws.max_row)
             paste_dataframe_to_sheet(ws, df_customer_master)
             
        ws_paste_n = nouhinsyo_wb["貼り付け用"]
        for r_idx, row in df_paste_sheet.iterrows():
            for c_idx, value in enumerate(row):
                ws_paste_n.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                
        # Bento for Nouhinsyo
        df_bento_for_nouhin = None
        if df_bento_sheet is not None:
             master_df = df_product_master.copy()
             if not master_df.empty and '商品名' in master_df.columns:
                 master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                 df_bento_for_nouhin = df_bento_sheet.copy()
                 df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                 df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]
        
        if df_bento_for_nouhin is not None and "注文弁当の抽出" in nouhinsyo_wb.sheetnames:
             safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin)
        
        if df_client_sheet is not None and "クライアント抽出" in nouhinsyo_wb.sheetnames:
             safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet)

        out_nouhin = io.BytesIO()
        nouhinsyo_wb.save(out_nouhin)
        b64_nouhin = base64.b64encode(out_nouhin.getvalue()).decode()
        
        return {
            "template_file": {
                "filename": f"{file.filename.replace('.pdf', '')}_数出表.xlsm",
                "data": b64_template
            },
            "nouhinsyo_file": {
                "filename": f"{file.filename.replace('.pdf', '')}_納品書.xlsx",
                "data": b64_nouhin
            }
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/masters/upload")
async def upload_master(file: UploadFile = File(...), type: str = "product"):
    try:
        content = await file.read()
        file_pattern = "商品マスタ" if type == "product" else "得意先マスタ"
        if type not in ["product", "customer"]:
             raise HTTPException(status_code=400, detail="Invalid type. Use 'product' or 'customer'.")
        
        # Check filename requirement (from original code)
        if type == "product" and "商品マスタ一覧" not in file.filename:
             raise HTTPException(status_code=400, detail="Filename must contain '商品マスタ一覧'")
        if type == "customer" and "得意先マスタ一覧" not in file.filename:
             raise HTTPException(status_code=400, detail="Filename must contain '得意先マスタ一覧'")

        success = save_master_file(ASSETS_DIR, content, file.filename, file_pattern)
        if success:
            return {"message": "File saved successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to save file")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/masters/info")
def get_master_info():
    """Get current master file names."""
    try:
        _, product_filename = load_master_csv(ASSETS_DIR, "商品マスタ")
        _, customer_filename = load_master_csv(ASSETS_DIR, "得意先マスタ")
        return {
            "product": product_filename or "未設定",
            "customer": customer_filename or "未設定"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Template file mappings
TEMPLATE_FILES = {
    "seal": ("seal.xlsx", "シールテンプレート"),
    "suudashiyo": ("template.xlsm", "数出表テンプレート"),
    "nouhinsyo": ("nouhinsyo.xlsx", "納品書テンプレート")
}

@app.get("/api/templates/info")
def get_template_info():
    """Get current template file info."""
    try:
        result = {}
        for key, (filename, label) in TEMPLATE_FILES.items():
            filepath = os.path.join(ASSETS_DIR, filename)
            if os.path.exists(filepath):
                mtime = os.path.getmtime(filepath)
                from datetime import datetime
                modified = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
                result[key] = {"filename": filename, "label": label, "exists": True, "modified": modified}
            else:
                result[key] = {"filename": filename, "label": label, "exists": False, "modified": None}
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/templates/upload")
async def upload_template(file: UploadFile = File(...), type: str = "seal"):
    """Upload a template file."""
    try:
        if type not in TEMPLATE_FILES:
            raise HTTPException(status_code=400, detail=f"Invalid template type. Use: {list(TEMPLATE_FILES.keys())}")
        
        expected_filename, label = TEMPLATE_FILES[type]
        content = await file.read()
        
        # Save template file
        save_path = os.path.join(ASSETS_DIR, expected_filename)
        with open(save_path, "wb") as f:
            f.write(content)
        
        return {"message": f"{label}を更新しました"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

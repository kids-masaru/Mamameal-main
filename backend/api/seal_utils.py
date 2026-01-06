from google import genai
from google.genai import types
import json
import os
from openpyxl import Workbook, load_workbook
import io

def generate_seal_data(pdf_bytes, model_name="gemini-3-flash-preview", api_key=None):
    """
    Generate seal data from PDF using Gemini (google-genai SDK).
    """
    if not api_key:
        raise ValueError("API Key is required for Gemini generation")

    client = genai.Client(api_key=api_key)
    
    seal_prompt = """
このPDFはシール表です。横4つ × 縦5つ(合計約20個)のブロックで構成されています。
各ブロックには以下の情報が含まれています:
1. クライアント名 (最上部): 小学校名または幼稚園名 + 「様」
2. 準備物 (クライアント名のすぐ下): パン箱入数、ご飯150gなど
3. クラス名 (中央、大きめの文字): チューリップ、さくらなど
4. 弁当数 (クラス名の下): 数値(例: 35、35+1)
5. 日付 (ブロック左下): MM/DD形式
6. 学年 (ブロック右下): 年長、年中など

以下のJSON形式で、全てのブロック情報を抽出してください:
{
  "blocks": [
    {
      "client_name": "博多南衆参コース様",
      "preparations": ["パン箱入数", "ご飯150g"],
      "class_name": "チューリップ",
      "meal_count": "35",
      "date": "12/10",
      "grade": "年長"
    }
  ]
}
重要: 全てのブロックを抽出してください。完全で有効なJSONのみを返してください。
"""
    try:
        response = client.models.generate_content(
            model=model_name,
            contents=[
                types.Content(
                    role="user",
                    parts=[
                        types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
                        types.Part.from_text(text=seal_prompt)
                    ]
                )
            ],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                max_output_tokens=65536
            )
        )
        
        text = response.text.strip()
        print(f"DEBUG: Gemini Raw Response (First 500 chars): {text[:500]}") # Log for debugging

        # Robust JSON extraction
        try:
            # First try direct parsing
            data = json.loads(text)
        except json.JSONDecodeError:
            # Fallback: finding the first { and last }
            import re
            match = re.search(r'(\{.*\})', text, re.DOTALL)
            if match:
                try:
                     # Try to clean up code block markers if stuck inside regex match
                     clean_json = match.group(1).replace("```json", "").replace("```", "")
                     data = json.loads(clean_json)
                except json.JSONDecodeError:
                     # Second fallback: try to fix common trailing comma issues (simple heuristic)
                     fixed_text = re.sub(r',\s*([\]}])', r'\1', match.group(1))
                     data = json.loads(fixed_text)
            else:
                 raise ValueError("No JSON object found in response")

        blocks = data if isinstance(data, list) else data.get('blocks', [])
        return blocks
        
    except Exception as e:
        print(f"Error in Gemini generation: {e}")
        # Print a snippet of the text if available to help debug
        if 'text' in locals():
             print(f"Failed JSON text snippet: {text[-500:]}")
        raise e

def create_seal_excel(blocks):
    """
    Create Excel file from seal blocks using template.
    Writes data to the 'Gemini抽出データ' sheet in the template.
    """
    # Get template path from environment or use default
    assets_dir = os.getenv("ASSETS_DIR", os.path.join(os.path.dirname(__file__), 'assets'))
    template_path = os.path.join(assets_dir, 'seal.xlsx')
    
    if os.path.exists(template_path):
        try:
            wb = load_workbook(template_path)
            # Try to get the data sheet, create if not exists
            if 'Gemini抽出データ' in wb.sheetnames:
                ws = wb['Gemini抽出データ']
                # Clear existing data (keep structure)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        if not isinstance(cell, type(None)):
                            try:
                                cell.value = None
                            except:
                                pass
            else:
                # Create new sheet for data
                ws = wb.create_sheet('Gemini抽出データ', 0)
            
            # Write headers
            headers = ['クライアント名', 'クラス名', '準備物', '弁当数', '日付', '学年']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data starting from row 2
            for idx, block in enumerate(blocks):
                row = idx + 2
                prep = block.get('preparations', [])
                prep_text = ', '.join(prep) if isinstance(prep, list) else str(prep)
                ws.cell(row=row, column=1, value=block.get('client_name', ''))
                ws.cell(row=row, column=2, value=block.get('class_name', ''))
                ws.cell(row=row, column=3, value=prep_text)
                ws.cell(row=row, column=4, value=block.get('meal_count', ''))
                ws.cell(row=row, column=5, value=block.get('date', ''))
                ws.cell(row=row, column=6, value=block.get('grade', ''))
            
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return out
            
        except Exception as e:
            print(f"Error using template: {e}")
            import traceback
            traceback.print_exc()
    
    # Fallback: create new workbook if template fails
    print("Falling back to new workbook")
    wb = Workbook()
    ws = wb.active
    ws.title = "シールデータ"
    
    headers = ['クライアント名', 'クラス名', '準備物', '弁当数', '日付', '学年']
    ws.append(headers)
    
    for block in blocks:
        prep = block.get('preparations', [])
        prep_text = ', '.join(prep) if isinstance(prep, list) else str(prep)
        ws.append([
            block.get('client_name', ''),
            block.get('class_name', ''),
            prep_text,
            block.get('meal_count', ''),
            block.get('date', ''),
            block.get('grade', '')
        ])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

